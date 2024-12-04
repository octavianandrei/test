import os
import pandas as pd
from openpyxl import load_workbook, Workbook

def is_excel_file(filename):
    try:
        load_workbook(filename)
        return True
    except Exception:
        return False

def sheet_exists(filename, sheet_name):
    try:
        book = load_workbook(filename)
        return sheet_name in book.sheetnames
    except Exception:
        return False

def delete_sheet_if_exists(filename, sheet_name):
    if os.path.exists(filename) and is_excel_file(filename):
        book = load_workbook(filename)
        if sheet_name in book.sheetnames:
            print(f"Sheet {sheet_name} exists. Deleting sheet.")
            std = book[sheet_name]
            book.remove(std)
            # Ensure at least one sheet is visible
            if not book.sheetnames:
                book.create_sheet('TempSheet')
            # Save the workbook after removing the sheet
            book.save(filename)
        book.close()  # Explicitly close the workbook after modifications

def create_empty_workbook(filename):
    wb = Workbook()
    temp_sheet = wb.active
    temp_sheet.title = "TempSheet"
    wb.save(filename)
    wb.close()
    file_path = filename
    os.chmod(file_path, 0o777)

def write_to_excel(filename, data, sheet_name, ensure_columns=None):
    # Check if data is empty and handle accordingly
    if len(data) == 0:
        print("No data to write.")
        return

    df = pd.DataFrame(data[1:], columns=data[0])

    # Ensure columns if specified
    if ensure_columns:
        for column in ensure_columns:
            if column not in df.columns:
                df[column] = None

    # Ensure the file exists
    if not os.path.exists(filename):
        print(f"File {filename} does not exist. Creating file.")
        create_empty_workbook(filename)

    if os.path.exists(filename) and is_excel_file(filename):
        book = load_workbook(filename)
        # Remove the temporary sheet if other sheets exist
        if "TempSheet" in book.sheetnames and len(book.sheetnames) > 1:
            std = book["TempSheet"]
            book.remove(std)
            book.save(filename)
            book.close()  # Always close after making modifications

        # Delete sheet if it exists before writing new data
        delete_sheet_if_exists(filename, sheet_name)

        # Append or write new data
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            writer._book = book
            writer._sheets = {ws.title: ws for ws in writer.book.worksheets}
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # If file doesn't exist, create new Excel file with the data
        with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def create_initial_excel_file(file_path, sheet_name):
    df = pd.DataFrame(columns=['Pipeline', 'Steps', 'Completion Weight', 'Median Completion Weight'])
    write_to_excel(file_path, [['Pipeline', 'Steps', 'Completion Weight', 'Median Completion Weight']] + df.values.tolist(), sheet_name, ensure_columns=['Pipeline', 'Steps', 'Completion Weight', 'Median Completion Weight'])
